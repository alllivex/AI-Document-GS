from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Sequence

import pandas as pd
from openpyxl import load_workbook

from app.engine.excel_utils import build_column_index_map, build_excel_column_letter_map
from app.models.template_models import FieldDefinition


@dataclass(frozen=True)
class LoadedTable:
    table_name: str
    source_file: str
    source_file_path: str
    dataframe: pd.DataFrame
    columns: list[str]
    row_count: int
    column_count: int
    column_index_map: dict[str, int]
    excel_column_letter_map: dict[str, str]


def load_data_tables(
    data_dir: Path,
    table_names: Sequence[str],
    fields: Sequence[FieldDefinition] | None = None,
) -> dict[str, LoadedTable]:
    fields_by_table: dict[str, list[FieldDefinition]] = {}
    for field in fields or []:
        fields_by_table.setdefault(field.table_name, []).append(field)
    return {
        table_name: load_data_table(data_dir, table_name, fields_by_table.get(table_name))
        for table_name in table_names
    }


def load_data_table(
    data_dir: Path,
    table_name: str,
    fields: Sequence[FieldDefinition] | None = None,
) -> LoadedTable:
    source_path = data_dir / f"{table_name}.xlsx"
    if not source_path.exists():
        raise FileNotFoundError(f"data table file not found for table '{table_name}': {source_path}")

    _validate_header_columns(_read_header_columns(source_path), table_name, source_path)
    dataframe = pd.read_excel(source_path, sheet_name=0)
    dataframe = _normalize_columns(dataframe, table_name, fields)
    _validate_dataframe(dataframe, table_name, source_path)

    columns = list(dataframe.columns)
    return LoadedTable(
        table_name=table_name,
        source_file=source_path.name,
        source_file_path=str(source_path),
        dataframe=dataframe,
        columns=columns,
        row_count=len(dataframe.index),
        column_count=len(columns),
        column_index_map=build_column_index_map(columns),
        excel_column_letter_map=build_excel_column_letter_map(columns),
    )


def _read_header_columns(source_path: Path) -> list[str]:
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        return [_clean_column_name(value) for value in first_row]
    finally:
        workbook.close()


def _validate_header_columns(columns: list[str], table_name: str, source_path: Path) -> None:
    if not columns or all(_is_missing_header(column) for column in columns):
        raise ValueError(f"data table '{table_name}' has no header row: {source_path}")

    invalid_columns = [
        index
        for index, column in enumerate(columns)
        if _is_missing_header(column)
    ]
    if invalid_columns:
        labels = ", ".join(str(index) for index in invalid_columns)
        raise ValueError(f"data table '{table_name}' has missing header columns at indexes: {labels}")

    duplicates = sorted({column for column in columns if columns.count(column) > 1})
    if duplicates:
        raise ValueError(f"data table '{table_name}' has duplicate header columns: {', '.join(duplicates)}")


def _normalize_columns(
    dataframe: pd.DataFrame,
    table_name: str,
    fields: Sequence[FieldDefinition] | None = None,
) -> pd.DataFrame:
    normalized = dataframe.copy()
    original_columns = [_clean_column_name(column) for column in normalized.columns]
    header_aliases = build_field_header_aliases(table_name, fields or [])
    columns = [header_aliases.get(column, column) for column in original_columns]

    duplicates = sorted({column for column in columns if columns.count(column) > 1})
    if duplicates:
        conflicts = [
            f"{canonical} <- {', '.join(original for original, mapped in zip(original_columns, columns) if mapped == canonical)}"
            for canonical in duplicates
        ]
        raise ValueError(
            f"data table '{table_name}' has duplicate header columns after schema mapping: {'; '.join(conflicts)}"
        )

    normalized.columns = columns
    return normalized


def build_field_header_aliases(
    table_name: str,
    fields: Sequence[FieldDefinition],
) -> dict[str, str]:
    alias_owners: dict[str, set[str]] = {}
    for field in fields:
        if field.table_name != table_name:
            continue
        alias_owners.setdefault(field.field_name, set()).add(field.field_name)
        chinese_name = field.field_name_cn.strip()
        if chinese_name:
            alias_owners.setdefault(chinese_name, set()).add(field.field_name)

    ambiguous = {
        alias: sorted(owners)
        for alias, owners in alias_owners.items()
        if len(owners) > 1
    }
    if ambiguous:
        details = "; ".join(
            f"{alias} -> {', '.join(owners)}"
            for alias, owners in sorted(ambiguous.items())
        )
        raise ValueError(f"entity schema has ambiguous field header aliases for table '{table_name}': {details}")

    return {alias: next(iter(owners)) for alias, owners in alias_owners.items()}


def _validate_dataframe(dataframe: pd.DataFrame, table_name: str, source_path: Path) -> None:
    if dataframe.shape[1] == 0:
        raise ValueError(f"data table '{table_name}' is empty and has no header row: {source_path}")

    invalid_columns = [
        index
        for index, column in enumerate(dataframe.columns)
        if _is_missing_header(column)
    ]
    if len(invalid_columns) == len(dataframe.columns):
        raise ValueError(f"data table '{table_name}' has no header row: {source_path}")
    if invalid_columns:
        labels = ", ".join(str(index) for index in invalid_columns)
        raise ValueError(f"data table '{table_name}' has missing header columns at indexes: {labels}")

    if dataframe.empty:
        raise ValueError(f"data table '{table_name}' is empty: {source_path}")


def _clean_column_name(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def _is_missing_header(column: str) -> bool:
    return not column or column.startswith("Unnamed:")
