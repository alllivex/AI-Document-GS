from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any
from uuid import uuid4

from jinja2 import Environment, StrictUndefined
from openpyxl import load_workbook

from app.engine.xlsx_template_parser import SIMPLE_VARIABLE_PATTERN


@dataclass(frozen=True)
class RenderedCellRun:
    text: str
    variable_path: str | None = None


@dataclass(frozen=True)
class RenderedCellBinding:
    coordinate: str
    rendered_text: str
    runs: tuple[RenderedCellRun, ...]


@dataclass(frozen=True)
class RenderXlsxResult:
    success: bool
    output_path: Path
    cells: tuple[RenderedCellBinding, ...] = ()
    error_message: str = ""


def render_xlsx_template(template_path: Path, context: dict[str, Any], output_path: Path) -> RenderXlsxResult:
    if not template_path.exists():
        return RenderXlsxResult(False, output_path, error_message=f"Template file not found: {template_path}")

    temp_output = output_path.with_name(f".{output_path.name}.{uuid4().hex}.tmp.xlsx")
    environment = Environment(undefined=StrictUndefined, autoescape=False)
    workbook = None
    try:
        workbook = load_workbook(template_path, data_only=False)
        worksheet = workbook.worksheets[0]
        bindings: list[RenderedCellBinding] = []
        for row in worksheet.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str) or "{{" not in value:
                    continue
                runs: list[RenderedCellRun] = []
                cursor = 0
                for match in SIMPLE_VARIABLE_PATTERN.finditer(value):
                    if match.start() > cursor:
                        runs.append(RenderedCellRun(value[cursor:match.start()]))
                    expression = match.group(0)
                    variable_path = f"{match.group(1)}.{match.group(2)}"
                    rendered_part = environment.from_string(expression).render(context)
                    runs.append(RenderedCellRun(rendered_part, variable_path))
                    cursor = match.end()
                if cursor < len(value):
                    runs.append(RenderedCellRun(value[cursor:]))
                rendered = "".join(item.text for item in runs)
                cell.value = rendered
                bindings.append(RenderedCellBinding(cell.coordinate, rendered, tuple(runs)))

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(temp_output)
        workbook.close()
        load_workbook(temp_output, read_only=True, data_only=False).close()
        temp_output.replace(output_path)
        return RenderXlsxResult(True, output_path, tuple(bindings))
    except Exception as exc:
        if workbook is not None:
            workbook.close()
        temp_output.unlink(missing_ok=True)
        output_path.unlink(missing_ok=True)
        return RenderXlsxResult(False, output_path, error_message=f"Failed to render xlsx template '{template_path}': {exc}")
