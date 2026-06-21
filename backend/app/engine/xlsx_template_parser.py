from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re

from jinja2 import Environment, StrictUndefined, TemplateSyntaxError
from openpyxl import load_workbook

from app.engine.template_canonicalizer import MissingTemplateVariable, TemplateVariableResolver, canonicalize_jinja_text
from app.engine.template_parser import TemplateFieldReference, extract_template_field_references_from_text
from app.models.template_models import FieldDefinition

JINJA_MARKER_PATTERN = re.compile(r"{[{%#]")
CONTROL_TAG_PATTERN = re.compile(r"{%|{#")
SIMPLE_VARIABLE_PATTERN = re.compile(
    r"{{\s*([^\W\d]\w*)\s*\.\s*([^\W\d]\w*)\s*}}",
    re.UNICODE,
)


@dataclass(frozen=True)
class XlsxCellTemplate:
    coordinate: str
    original_text: str
    canonical_text: str
    variable_paths: tuple[str, ...]


@dataclass(frozen=True)
class XlsxTemplateAnalysis:
    sheet_name: str
    cells: tuple[XlsxCellTemplate, ...]
    references: tuple[TemplateFieldReference, ...]
    original_var_paths_by_canonical: dict[str, str]
    missing_variables: tuple[MissingTemplateVariable, ...]
    issues: tuple[tuple[str, str, str | None], ...]


def analyze_xlsx_template(template_path: Path, fields: list[FieldDefinition]) -> XlsxTemplateAnalysis:
    workbook = load_workbook(template_path, data_only=False)
    if not workbook.worksheets:
        raise ValueError("XLSX template must contain at least one worksheet")

    worksheet = workbook.worksheets[0]
    resolver = TemplateVariableResolver(fields)
    cells: list[XlsxCellTemplate] = []
    references: dict[tuple[str, str], TemplateFieldReference] = {}
    original_by_canonical: dict[str, str] = {}
    missing: list[MissingTemplateVariable] = []
    issues: list[tuple[str, str, str | None]] = []
    environment = Environment(undefined=StrictUndefined, autoescape=False)

    for other_sheet in workbook.worksheets[1:]:
        for row in other_sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and JINJA_MARKER_PATTERN.search(cell.value):
                    issues.append((
                        "xlsx_variable_outside_first_sheet",
                        f"Jinja marker found outside the first worksheet: {other_sheet.title}!{cell.coordinate}",
                        cell.coordinate,
                    ))

    for row in worksheet.iter_rows():
        for cell in row:
            value = cell.value
            if not isinstance(value, str) or not JINJA_MARKER_PATTERN.search(value):
                continue
            if cell.data_type == "f" or value.startswith("="):
                issues.append(("xlsx_formula_variable_unsupported", f"Formula cell contains Jinja: {cell.coordinate}", cell.coordinate))
                continue
            if CONTROL_TAG_PATTERN.search(value):
                issues.append(("xlsx_unsupported_jinja_statement", f"Control statements are not supported: {cell.coordinate}", cell.coordinate))
                continue
            try:
                environment.parse(value)
            except TemplateSyntaxError as exc:
                issues.append(("xlsx_jinja_syntax_invalid", f"Invalid Jinja syntax at {cell.coordinate}: {exc}", cell.coordinate))
                continue

            canonicalized = canonicalize_jinja_text(value, resolver)
            original_by_canonical.update(canonicalized.original_var_paths_by_canonical)
            missing.extend(canonicalized.missing_variables)
            cell_refs = extract_template_field_references_from_text(canonicalized.text)
            for reference in cell_refs:
                references[(reference.table_name, reference.field_name)] = reference

            variable_paths = tuple(f"{table}.{field}" for table, field in SIMPLE_VARIABLE_PATTERN.findall(canonicalized.text))
            expression_count = canonicalized.text.count("{{")
            if len(variable_paths) != expression_count:
                issues.append((
                    "xlsx_unsupported_jinja_statement",
                    f"Only simple table.field expressions are supported: {cell.coordinate}",
                    cell.coordinate,
                ))
            cells.append(XlsxCellTemplate(cell.coordinate, value, canonicalized.text, variable_paths))

    merged_ranges = list(worksheet.merged_cells.ranges)
    for cell_template in cells:
        for merged_range in merged_ranges:
            if cell_template.coordinate in merged_range and cell_template.coordinate != merged_range.start_cell.coordinate:
                issues.append((
                    "xlsx_merged_cell_variable_invalid",
                    f"Jinja must be placed in the top-left merged cell: {cell_template.coordinate}",
                    cell_template.coordinate,
                ))

    unique_missing = {(item.original_var_path, item.reason): item for item in missing}
    result = XlsxTemplateAnalysis(
        sheet_name=worksheet.title,
        cells=tuple(cells),
        references=tuple(references.values()),
        original_var_paths_by_canonical=original_by_canonical,
        missing_variables=tuple(unique_missing.values()),
        issues=tuple(issues),
    )
    workbook.close()
    return result
