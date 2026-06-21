from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable
from uuid import uuid4

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Color

from app.engine.trace_builder import BuildTracePreviewInput, BuildTracePreviewResult, build_trace_file
from app.engine.xlsx_renderer import RenderedCellBinding
from app.models.preview_models import PreviewFile, PreviewRun, PreviewRunStyle, PreviewTableBlock, PreviewTableCell
from app.models.trace_models import TraceItem


def build_xlsx_trace_and_preview(
    input_data: BuildTracePreviewInput,
    rendered_cells: tuple[RenderedCellBinding, ...],
) -> BuildTracePreviewResult:
    trace_by_path = input_data.trace_map
    used_paths = list(dict.fromkeys(
        run.variable_path for cell in rendered_cells for run in cell.runs if run.variable_path
    ))
    trace_items = _unique_trace_items(
        item for path in used_paths for item in trace_by_path.get(path, [])[:1]
    )
    trace_path = input_data.output_path.with_suffix(".trace.json")
    preview_path = input_data.output_path.with_suffix(".preview.json")
    _write_model(trace_path, build_trace_file(input_data, trace_items, [], [], []))
    _write_model(preview_path, _build_xlsx_preview(input_data, rendered_cells, trace_by_path))
    return BuildTracePreviewResult(
        trace_file_path=trace_path,
        preview_file_path=preview_path,
        trace_count=len(trace_items),
        ai_block_count=0,
    )


def _build_xlsx_preview(
    input_data: BuildTracePreviewInput,
    rendered_cells: tuple[RenderedCellBinding, ...],
    trace_by_path: dict[str, list[TraceItem]],
) -> PreviewFile:
    workbook = load_workbook(input_data.output_path, data_only=False)
    worksheet = workbook.worksheets[0]
    binding_by_coordinate = {item.coordinate: item for item in rendered_cells}
    merged_by_coordinate: dict[str, tuple[int, int]] = {}
    covered_coordinates: set[str] = set()
    for merged_range in worksheet.merged_cells.ranges:
        merged_by_coordinate[merged_range.start_cell.coordinate] = (
            merged_range.max_row - merged_range.min_row + 1,
            merged_range.max_col - merged_range.min_col + 1,
        )
        for row in worksheet.iter_rows(
            min_row=merged_range.min_row,
            max_row=merged_range.max_row,
            min_col=merged_range.min_col,
            max_col=merged_range.max_col,
        ):
            for cell in row:
                if cell.coordinate != merged_range.start_cell.coordinate:
                    covered_coordinates.add(cell.coordinate)

    populated = [cell for row in worksheet.iter_rows() for cell in row if cell.value is not None]
    max_row = max((cell.row for cell in populated), default=0)
    max_column = max((cell.column for cell in populated), default=0)
    rows: list[list[PreviewTableCell]] = []
    worksheet_rows = (
        worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_column)
        if max_row and max_column
        else []
    )
    for row in worksheet_rows:
        preview_row: list[PreviewTableCell] = []
        for cell in row:
            if isinstance(cell, MergedCell) or cell.coordinate in covered_coordinates:
                continue
            binding = binding_by_coordinate.get(cell.coordinate)
            runs = _preview_runs(binding, trace_by_path) if binding else None
            rowspan, colspan = merged_by_coordinate.get(cell.coordinate, (1, 1))
            traced_runs = [run for run in (runs or []) if run.trace_id]
            preview_row.append(
                PreviewTableCell(
                    text=_cell_text(cell.value),
                    trace_id=traced_runs[0].trace_id if len(runs or []) == 1 and traced_runs else None,
                    trace_kind="field" if len(runs or []) == 1 and traced_runs else None,
                    runs=runs,
                    rowspan=rowspan,
                    colspan=colspan,
                    style=_cell_style(cell),
                    width_px=_column_width_px(worksheet.column_dimensions[cell.column_letter].width),
                )
            )
        rows.append(preview_row)
    workbook.close()
    blocks = [PreviewTableBlock(block_id=f"xlsx_table_{uuid4().hex}", headers=[], rows=rows)] if rows else []
    return PreviewFile(
        doc_id=input_data.doc_id,
        task_id=input_data.task_id,
        title=input_data.output_file,
        output_file=input_data.output_file,
        primary_key_value=input_data.primary_key_value,
        blocks=blocks,
        created_at=datetime.now(timezone.utc),
    )


def _preview_runs(binding: RenderedCellBinding, trace_by_path: dict[str, list[TraceItem]]) -> list[PreviewRun]:
    runs: list[PreviewRun] = []
    for run in binding.runs:
        matching = trace_by_path.get(run.variable_path or "", [])
        trace_item = matching[0] if matching else None
        runs.append(PreviewRun(text=run.text, trace_id=trace_item.trace_id if trace_item else None, trace_kind="field" if trace_item else None))
    return runs


def _cell_style(cell: Any) -> PreviewRunStyle | None:
    style = PreviewRunStyle(
        bold=bool(cell.font.bold) or None,
        italic=bool(cell.font.italic) or None,
        underline=bool(cell.font.underline) or None,
        alignment=cell.alignment.horizontal if cell.alignment.horizontal in {"left", "center", "right", "justify"} else None,
        color=_color_value(cell.font.color),
        background_color=_color_value(cell.fill.fgColor),
        font_size_pt=float(cell.font.sz) if cell.font.sz else None,
    )
    return style if any(value is not None for value in style.model_dump().values()) else None


def _color_value(color: Color | None) -> str | None:
    if color is None or color.type != "rgb" or not color.rgb:
        return None
    value = str(color.rgb)
    if len(value) == 8 and value[:2].upper() == "00":
        return None
    return f"#{value[-6:]}"


def _column_width_px(width: float | None) -> float | None:
    return round(width * 7 + 5, 1) if width else None


def _cell_text(value: Any) -> str:
    return "" if value is None else str(value)


def _unique_trace_items(items: Iterable[TraceItem]) -> list[TraceItem]:
    result: list[TraceItem] = []
    seen: set[str] = set()
    for item in items:
        if item.trace_id not in seen:
            seen.add(item.trace_id)
            result.append(item)
    return result


def _write_model(path: Path, model: Any) -> None:
    path.write_text(json.dumps(model.model_dump(mode="json"), ensure_ascii=False, indent=2), encoding="utf-8")
