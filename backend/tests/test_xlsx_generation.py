from __future__ import annotations

import json
from pathlib import Path
from copy import copy

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Color

from app.engine.generation_runner import GenerateTaskInput, GenerationDependencies, generate_task
from app.engine.xlsx_renderer import render_xlsx_template
from app.engine.xlsx_preview_builder import _color_value
from app.engine.xlsx_template_canonicalizer import canonicalize_xlsx_template
from app.engine.xlsx_template_parser import analyze_xlsx_template
from app.models.enums import DataType, TaskStatus
from app.models.template_models import FieldDefinition
from app.engine.value_formatter import TraceValue

from test_generation_runner import configure_chinese_field_names, setup_project, write_chinese_header_excel_data


FIELDS = [
    FieldDefinition(
        table_name="customer_info",
        table_name_cn="客户信息表",
        field_name="customer_name",
        field_name_cn="客户名称",
        data_type=DataType.STRING,
    ),
    FieldDefinition(
        table_name="loan_summary",
        table_name_cn="贷款汇总表",
        field_name="loan_balance",
        field_name_cn="贷款余额",
        data_type=DataType.AMOUNT,
    ),
]


def test_xlsx_preview_treats_transparent_default_color_as_no_color() -> None:
    assert _color_value(Color(rgb="00000000")) is None
    assert _color_value(Color(rgb="FFFF0000")) == "#FF0000"


def write_xlsx_template(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Report"
    worksheet["A1"] = "客户"
    worksheet["B1"] = "{{ customer_info.customer_name }}"
    worksheet["A2"] = "余额"
    worksheet["B2"] = "金额：{{ loan_summary.loan_balance }}"
    bold_font = copy(worksheet["A1"].font)
    bold_font.bold = True
    worksheet["A1"].font = bold_font
    worksheet.column_dimensions["B"].width = 24
    workbook.save(path)


def test_xlsx_parser_and_canonicalizer_support_chinese_aliases(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "canonical.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "{{ 客户信息表.客户名称 }}"
    workbook.save(template_path)

    analysis = analyze_xlsx_template(template_path, FIELDS)
    result = canonicalize_xlsx_template(template_path, output_path, FIELDS)

    assert not analysis.issues
    assert analysis.references[0].variable_path == "customer_info.customer_name"
    assert result.original_var_paths_by_canonical["customer_info.customer_name"] == "客户信息表.客户名称"
    assert load_workbook(output_path).active["A1"].value == "{{ customer_info.customer_name }}"
    assert load_workbook(template_path).active["A1"].value == "{{ 客户信息表.客户名称 }}"


def test_xlsx_parser_reports_unsupported_statements_and_other_sheet_variables(tmp_path: Path) -> None:
    template_path = tmp_path / "invalid.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "{% if customer_info.customer_name %}yes{% endif %}"
    workbook.create_sheet("Other")["B2"] = "{{ customer_info.customer_name }}"
    workbook.save(template_path)

    analysis = analyze_xlsx_template(template_path, FIELDS)
    assert {issue[0] for issue in analysis.issues} == {
        "xlsx_unsupported_jinja_statement",
        "xlsx_variable_outside_first_sheet",
    }


def test_xlsx_renderer_preserves_layout_and_renders_mixed_text(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    write_xlsx_template(template_path)
    result = render_xlsx_template(
        template_path,
        {
            "customer_info": {"customer_name": TraceValue(raw_value="Acme", display_value="Acme", trace_id="t1")},
            "loan_summary": {"loan_balance": TraceValue(raw_value=1200, display_value="1,200.00", trace_id="t2")},
        },
        output_path,
    )

    assert result.success
    workbook = load_workbook(output_path)
    assert workbook.active["B1"].value == "Acme"
    assert workbook.active["B2"].value == "金额：1,200.00"
    assert workbook.active.column_dimensions["B"].width == 24
    assert workbook.active["A1"].font.bold is True


def test_generate_task_creates_xlsx_trace_and_preview_for_each_main_row(tmp_path: Path) -> None:
    connection, settings = setup_project(tmp_path)
    template_path = settings.templates_dir / "due_diligence.xlsx"
    write_xlsx_template(template_path)
    (settings.templates_dir / "due_diligence.docx").unlink()
    connection.execute(
        "UPDATE templates SET template_file = ?, template_path = ?, ai_enabled_default = 0 WHERE id = 1",
        ("due_diligence.xlsx", "templates/due_diligence.xlsx"),
    )
    connection.commit()

    result = generate_task(
        GenerateTaskInput(task_id="task_001", ai_enabled=True),
        connection=connection,
        settings=settings,
        dependencies=GenerationDependencies(),
    )

    assert result.status == TaskStatus.COMPLETED
    assert result.success_count == 3
    output_dir = settings.tasks_dir / "task_001" / "output"
    outputs = sorted(output_dir.glob("*.xlsx"))
    assert len(outputs) == 3
    assert load_workbook(outputs[0]).active["B1"].value == "Acme"

    preview = json.loads(sorted(output_dir.glob("*.preview.json"))[0].read_text(encoding="utf-8"))
    traced_runs = [
        run
        for block in preview["blocks"]
        for row in block.get("rows", [])
        for cell in row
        for run in cell.get("runs") or []
        if run.get("trace_id")
    ]
    assert {run["text"] for run in traced_runs} == {"Acme", "1200.00"}

    trace = json.loads(sorted(output_dir.glob("*.trace.json"))[0].read_text(encoding="utf-8"))
    assert {item["var_path"] for item in trace["trace_items"]} == {
        "customer_info.customer_name",
        "loan_summary.loan_balance",
    }
    records = connection.execute("SELECT output_filename, ai_status, ai_block_count FROM documents").fetchall()
    assert all(row["output_filename"].endswith(".xlsx") for row in records)
    assert {row["ai_status"] for row in records} == {"not_used"}
    assert {row["ai_block_count"] for row in records} == {0}


def test_generate_xlsx_from_chinese_business_data_headers(tmp_path: Path) -> None:
    connection, settings = setup_project(tmp_path)
    configure_chinese_field_names(connection)
    write_chinese_header_excel_data(settings)
    template_path = settings.templates_dir / "due_diligence.xlsx"
    write_xlsx_template(template_path)
    (settings.templates_dir / "due_diligence.docx").unlink()
    connection.execute(
        "UPDATE templates SET template_file = ?, template_path = ? WHERE id = 1",
        ("due_diligence.xlsx", "templates/due_diligence.xlsx"),
    )
    connection.commit()

    result = generate_task(
        GenerateTaskInput(task_id="task_001", ai_enabled=False),
        connection=connection,
        settings=settings,
        dependencies=GenerationDependencies(),
    )

    assert result.status == TaskStatus.COMPLETED
    output = next((settings.tasks_dir / "task_001" / "output").glob("*.xlsx"))
    assert load_workbook(output).active["B1"].value == "Acme"
